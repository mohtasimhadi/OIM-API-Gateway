import os
import shutil
from fastapi import UploadFile
import pandas as pd
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl import Workbook

UPLOAD_DIRECTORY = "./uploads"
os.makedirs(UPLOAD_DIRECTORY, exist_ok=True)

def save_upload_file(upload_file: UploadFile) -> str:
    file_path = os.path.join(UPLOAD_DIRECTORY, upload_file.filename)
    with open(file_path, "wb") as f:
        shutil.copyfileobj(upload_file.file, f)
    return file_path

def remove_file(file_path: str):
    if os.path.exists(file_path):
        os.remove(file_path)

def create_xlsx(data):
    df = pd.DataFrame(data['analysis']['track_data'])
    df = df[['track_id', 'confidence', 'area', 'perimeter', 'circularity', 'eccentricity']]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def style_cell(cell):
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="right", vertical="center")

    wb = Workbook()
    ws = wb.active

    # Basic metadata
    ws['A1'] = "Bed No"
    ws['B1'] = data['bed_number']

    ws['A2'] = "Collection Date"
    ws['B2'] = data['collection_date']

    ws['A3'] = "Plants"
    ws['B3'] = ', '.join(data['plants'])

    # Styling metadata
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['A2'].font = Font(bold=True)
    ws['A3'].font = Font(bold=True)
    ws['A2'].alignment = ws['A3'].alignment = Alignment(horizontal="left")

    start_row_for_dataframe = 7

    # Add dataframe data
    rows = [df.columns.tolist()] + df.values.tolist()

    for row_idx, row_data in enumerate(rows, start=start_row_for_dataframe):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == start_row_for_dataframe:
                style_cell(cell)
            else:
                cell.border = border

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    final_row = start_row_for_dataframe + len(rows)

    # Add averages
    ws[f'A{final_row}'] = "Average"
    ws[f'B{final_row}'] = df['confidence'].mean()
    ws[f'C{final_row}'] = df['area'].mean()
    ws[f'D{final_row}'] = df['perimeter'].mean()
    ws[f'E{final_row}'] = df['circularity'].mean()
    ws[f'F{final_row}'] = df['eccentricity'].mean()

    for col in 'ABCDEF':
        ws[f'{col}{final_row}'].font = Font(bold=True, italic=True)

    # Create charts
    def create_chart(ws, title, y_axis, values_refs, chart_position):
        chart = LineChart()
        chart.title = title
        chart.y_axis.title = y_axis
        chart.x_axis.title = "Track ID"
        chart.style = 13
        chart.layout = Layout(ManualLayout(x=0.25, y=0.25, w=0.9, h=0.7))

        for values_ref in values_refs:
            series = Series(values_ref, title_from_data=True)
            chart.append(series)

        # Add the chart to the worksheet
        ws.add_chart(chart, chart_position)

    # Define data range for charts
    categories = Reference(ws, min_col=1, min_row=start_row_for_dataframe + 1, max_row=final_row - 1)
    
    confidence_values = Reference(ws, min_col=2, min_row=start_row_for_dataframe, max_row=final_row - 1)
    area_values = Reference(ws, min_col=3, min_row=start_row_for_dataframe, max_row=final_row - 1)
    perimeter_values = Reference(ws, min_col=4, min_row=start_row_for_dataframe, max_row=final_row - 1)
    circularity_values = Reference(ws, min_col=5, min_row=start_row_for_dataframe, max_row=final_row - 1)
    eccentricity_values = Reference(ws, min_col=6, min_row=start_row_for_dataframe, max_row=final_row - 1)

    create_chart(ws, "Confidence, Circularity, Eccentricity", "Percentage (%)", 
                 [confidence_values, circularity_values, eccentricity_values], "H5")

    create_chart(ws, "Area Perimeter", "Pixels", [perimeter_values], "H21")
    create_chart(ws, "Area", "Pixel^2", [area_values], "H37")

    return wb